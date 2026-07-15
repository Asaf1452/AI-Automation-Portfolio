"""
CSV Report Generator
--------------------
Turns a plain CSV file into a polished Excel report:

- Styled header row with filters
- Auto-sized columns
- Number formatting and a totals row for numeric columns
- Optional summary sheet with totals per group (e.g. sales per region)

Usage:
    python report.py sales.csv
    python report.py sales.csv --group-by Region
    python report.py sales.csv --output monthly_report.xlsx

Requires: Python 3.8+ and one library:  pip install openpyxl
"""

import argparse
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Colors used for styling (feel free to change)
HEADER_FILL = PatternFill("solid", fgColor="2F5496")  # dark blue
HEADER_FONT = Font(color="FFFFFF", bold=True)          # white, bold
TOTAL_FONT = Font(bold=True)


def read_csv(csv_path):
    """Read the CSV file.

    Returns (headers, rows) where headers is a list of column names
    and rows is a list of lists (one list per data row).
    """
    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        all_rows = [row for row in reader if row]  # skip blank lines

    if len(all_rows) < 2:
        raise SystemExit("Error: the CSV needs a header row and at least one data row.")

    headers, rows = all_rows[0], all_rows[1:]
    return headers, rows


def convert_numbers(rows):
    """Convert number-looking text ("42", "19.99") into real numbers
    so Excel can format and sum them. Everything else stays text.
    """
    converted = []
    for row in rows:
        new_row = []
        for value in row:
            text = value.strip().replace(",", "")  # allow "1,250"
            try:
                number = float(text)
                # Show 42 instead of 42.0 for whole numbers
                new_row.append(int(number) if number.is_integer() else number)
            except ValueError:
                new_row.append(value.strip())
        converted.append(new_row)
    return converted


def numeric_columns(headers, rows):
    """Find which columns contain only numbers.

    Returns a list of column indexes (0 = first column).
    """
    indexes = []
    for col in range(len(headers)):
        values = [row[col] for row in rows if col < len(row) and row[col] != ""]
        if values and all(isinstance(v, (int, float)) for v in values):
            indexes.append(col)
    return indexes


def style_header_row(sheet):
    """Make the first row look like a report header."""
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def autofit_columns(sheet):
    """Set each column's width to fit its longest value."""
    for column_cells in sheet.columns:
        longest = max(len(str(cell.value or "")) for cell in column_cells)
        letter = get_column_letter(column_cells[0].column)
        sheet.column_dimensions[letter].width = longest + 4


def add_data_sheet(workbook, headers, rows, number_cols):
    """Create the main sheet: the full table plus a totals row."""
    sheet = workbook.active
    sheet.title = "Data"

    sheet.append(headers)
    for row in rows:
        sheet.append(row)

    # Totals row: "TOTAL" under the first column, sums under numeric columns
    totals = ["TOTAL"] + [""] * (len(headers) - 1)
    for col in number_cols:
        if col == 0:
            continue  # keep the TOTAL label
        totals[col] = sum(r[col] for r in rows if isinstance(r[col], (int, float)))
    sheet.append(totals)
    for cell in sheet[sheet.max_row]:
        cell.font = TOTAL_FONT

    # Format numeric columns with thousand separators
    for col in number_cols:
        for row_num in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row_num, column=col + 1)
            if isinstance(cell.value, float):
                cell.number_format = "#,##0.00"
            elif isinstance(cell.value, int):
                cell.number_format = "#,##0"

    style_header_row(sheet)
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{sheet.max_row - 1}"
    sheet.freeze_panes = "A2"  # keep the header visible while scrolling
    autofit_columns(sheet)


def add_summary_sheet(workbook, headers, rows, number_cols, group_by):
    """Create a second sheet with totals per group.

    Example: --group-by Region gives one row per region
    with the sum of every numeric column.
    """
    if group_by not in headers:
        raise SystemExit(
            f"Error: column '{group_by}' not found. Available: {', '.join(headers)}"
        )

    group_col = headers.index(group_by)
    sum_cols = [c for c in number_cols if c != group_col]

    # Add up the numeric columns for each group value
    groups = {}
    for row in rows:
        key = row[group_col]
        if key not in groups:
            groups[key] = [0] * len(sum_cols)
        for i, col in enumerate(sum_cols):
            if isinstance(row[col], (int, float)):
                groups[key][i] += row[col]

    sheet = workbook.create_sheet(f"Summary by {group_by}")
    sheet.append([group_by] + [headers[c] for c in sum_cols])
    for key in sorted(groups):
        sheet.append([key] + groups[key])

    # Format the number cells
    for row_num in range(2, sheet.max_row + 1):
        for col_num in range(2, len(sum_cols) + 2):
            cell = sheet.cell(row=row_num, column=col_num)
            if isinstance(cell.value, float):
                cell.number_format = "#,##0.00"
            elif isinstance(cell.value, int):
                cell.number_format = "#,##0"

    style_header_row(sheet)
    autofit_columns(sheet)


def main():
    parser = argparse.ArgumentParser(
        description="Turn a CSV file into a formatted Excel report."
    )
    parser.add_argument("csv_file", help="Path of the CSV file to convert")
    parser.add_argument(
        "--group-by",
        metavar="COLUMN",
        help="Also create a summary sheet with totals per value of this column",
    )
    parser.add_argument(
        "--output",
        metavar="FILE",
        help="Name of the Excel file to create (default: <csv name>_report.xlsx)",
    )
    args = parser.parse_args()

    csv_path = Path(args.csv_file).expanduser().resolve()
    if not csv_path.is_file():
        raise SystemExit(f"Error: '{csv_path}' does not exist.")

    output = Path(args.output) if args.output else csv_path.with_name(
        csv_path.stem + "_report.xlsx"
    )

    headers, raw_rows = read_csv(csv_path)
    rows = convert_numbers(raw_rows)
    number_cols = numeric_columns(headers, rows)

    workbook = Workbook()
    add_data_sheet(workbook, headers, rows, number_cols)
    if args.group_by:
        add_summary_sheet(workbook, headers, rows, number_cols, args.group_by)

    workbook.save(output)
    print(f"Report created: {output}")
    print(f"  Rows: {len(rows)}  |  Columns: {len(headers)}", end="")
    if args.group_by:
        print(f"  |  Summary sheet: by {args.group_by}", end="")
    print()


if __name__ == "__main__":
    main()
